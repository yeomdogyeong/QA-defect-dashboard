"""
리포트 출력 (고품질 · 한글).

차트 4종(원형 2 + 추세선 1 + 막대 1)을 크게 배치.
- to_excel: 'Dashboard' 시트 = 제목 배너 + KPI 카드 + 2x2 차트 / 'Data' 시트 = 원본
- to_pdf:   1페이지 = 제목 + KPI 카드 + 2x2 차트 / 2페이지 = 모듈별 결함 표

한글은 맑은 고딕(Windows) 우선, 없으면 다른 한글 폰트로 자동 폴백.
"""

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from src import charts as C
from src import metrics as M

_NAVY = "1F2A44"

# 리포트에 실을 4개 차트 (원형 status·rca + 추세 trend + 막대 module)
_REPORT_CHARTS = ("status", "rca", "trend", "module")


# --------------------------------------------------------------------------- #
#  PDF용 한글 폰트 등록 (reportlab)
# --------------------------------------------------------------------------- #
def _register_pdf_fonts() -> tuple[str, str]:
    """한글 정체/볼드 TTF를 reportlab에 등록. (본문폰트, 볼드폰트) 반환."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    reg, bold = "Helvetica", "Helvetica-Bold"
    path = C.KOREAN_FONT_PATH
    if path:
        try:
            if path.lower().endswith(".ttc"):
                pdfmetrics.registerFont(TTFont("KFont", path, subfontIndex=0))
            else:
                pdfmetrics.registerFont(TTFont("KFont", path))
            reg, bold = "KFont", "KFont"
        except Exception:
            return "Helvetica", "Helvetica-Bold"
    bpath = C.KOREAN_FONT_BOLD_PATH
    if reg == "KFont" and bpath and not bpath.lower().endswith(".ttc"):
        try:
            pdfmetrics.registerFont(TTFont("KFontB", bpath))
            bold = "KFontB"
        except Exception:
            bold = "KFont"
    return reg, bold


# --------------------------------------------------------------------------- #
#  Excel
# --------------------------------------------------------------------------- #
def to_excel(df: pd.DataFrame, freq: str = "W") -> bytes:
    from src import labels as L

    imgs = C.all_charts(df, freq=freq)
    wb = Workbook()
    ws = wb.active
    ws.title = "대시보드"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:S2")
    t = ws["A1"]
    t.value = "  QA 결함 대시보드 리포트"
    t.fill = PatternFill("solid", fgColor=_NAVY)
    t.font = Font(bold=True, size=20, color="FFFFFF")
    t.alignment = Alignment(vertical="center")
    ws.merge_cells("A3:S3")
    ws["A3"].value = f"  생성 {datetime.now():%Y-%m-%d %H:%M}   |   결함 {len(df):,}건"
    ws["A3"].font = Font(size=10, color="868E96")

    def _place(png: bytes, anchor: str, w: int, h: int):
        img = XLImage(io.BytesIO(png))
        img.width, img.height = w, h
        ws.add_image(img, anchor)

    # KPI 카드(풀폭) + 2x2 차트 (간격 약 2셀로 붙임)
    _place(imgs["kpi"], "A5", 1150, 177)
    grid = [("status", "A16"), ("rca", "L16"), ("trend", "A37"), ("module", "L37")]
    for key, anchor in grid:
        _place(imgs[key], anchor, 560, 385)

    # 데이터 시트 (헤더·범주값 한글화)
    ws2 = wb.create_sheet("데이터")
    disp = df.copy()
    disp["status"] = disp["status"].map(lambda v: L.STATUS_KO.get(v, v))
    disp["severity"] = disp["severity"].map(lambda v: L.SEVERITY_KO.get(v, v))
    disp["priority"] = disp["priority"].map(lambda v: L.PRIORITY_KO.get(v, v))
    disp["found_phase"] = disp["found_phase"].map(lambda v: L.PHASE_KO.get(v, v))
    disp["root_cause"] = disp["root_cause"].map(lambda v: L.ROOT_CAUSE_KO.get(v, v))
    disp["module"] = disp["module"].map(lambda v: L.MODULE_KO.get(v, v))
    for en, ko in L.MODULE_KO.items():             # summary 안의 영문 모듈명도 치환
        disp["summary"] = disp["summary"].str.replace(f"[{en}]", f"[{ko}]", regex=False)
    disp = disp.rename(columns=L.COLUMN_KO)
    for r_idx, row in enumerate(dataframe_to_rows(disp, index=False, header=True)):
        for c_idx, val in enumerate(row, start=1):
            cell = ws2.cell(row=r_idx + 1, column=c_idx, value=val)
            if r_idx == 0:
                cell.fill = PatternFill("solid", fgColor=_NAVY)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    for col_cells in ws2.columns:
        letter = col_cells[0].column_letter
        width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws2.column_dimensions[letter].width = min(max(width + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
#  PDF
# --------------------------------------------------------------------------- #
def to_pdf(df: pd.DataFrame, freq: str = "W") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Image as RLImage, PageBreak, Paragraph, SimpleDocTemplate, Spacer,
        Table, TableStyle,
    )

    kfont, kbold = _register_pdf_fonts()
    ko = kfont == "KFont"

    imgs = C.all_charts(df, freq=freq, tall=True)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, title="QA Defect Report",
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.2 * cm, bottomMargin=1.0 * cm,
    )
    navy = colors.HexColor("#1F2A44")
    title_st = ParagraphStyle("t", fontName=kbold, textColor=navy, fontSize=22, leading=26)
    sub_st = ParagraphStyle("s", fontName=kfont, textColor=colors.HexColor("#868E96"),
                            fontSize=10, leading=14)
    h2 = ParagraphStyle("h2", fontName=kbold, textColor=navy, fontSize=15, leading=19,
                        spaceAfter=8)

    def rl_img(png: bytes, w_cm: float, h_cm: float) -> RLImage:
        return RLImage(io.BytesIO(png), width=w_cm * cm, height=h_cm * cm)

    t_title = "QA 결함 대시보드 리포트" if ko else "QA Defect Dashboard Report"
    t_sub = (f"생성 {datetime.now():%Y-%m-%d %H:%M}  &nbsp;|&nbsp;  결함 {len(df):,}건 분석"
             if ko else f"Generated {datetime.now():%Y-%m-%d %H:%M}  |  {len(df):,} defects")

    story = [
        Paragraph(t_title, title_st),
        Paragraph(t_sub, sub_st),
        Spacer(1, 0.35 * cm),
        rl_img(imgs["kpi"], 17.6, 2.71),      # KPI 카드 풀폭
        Spacer(1, 0.35 * cm),
    ]

    # 2x2 차트 그리드 (세로형 · A4 꽉 채우기). 이미지 비율 6:7과 일치시켜 왜곡 없음.
    w, h = 8.83, 10.3
    grid = Table(
        [[rl_img(imgs["status"], w, h), rl_img(imgs["rca"], w, h)],
         [rl_img(imgs["trend"], w, h), rl_img(imgs["module"], w, h)]],
        colWidths=[9.0 * cm, 9.0 * cm],
    )
    grid.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(grid)

    # ---- 2페이지: 모듈별 결함 표 ----
    story.append(PageBreak())
    story.append(Paragraph("모듈별 결함 건수" if ko else "Defects by Module", h2))
    mod = M.defect_density_by_module(df)
    header = ["모듈", "건수"] if ko else ["Module", "Count"]
    from src import labels as _L
    rows = [header] + [[_L.MODULE_KO.get(m, m), c] for m, c in mod.values.tolist()]
    tbl = Table(rows, colWidths=[10 * cm, 5 * cm], hAlign="LEFT")
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), kbold),
        ("FONTNAME", (0, 1), (-1, -1), kfont),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CED4DA")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F3F9")]),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(tbl)

    doc.build(story)
    return buf.getvalue()
